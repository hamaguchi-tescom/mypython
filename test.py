import os
import locale
import glob
import datetime
import codecs
import sys
from pathlib import Path
from openpyxl import Workbook
import openpyxl


# # プログラム実行時の年月日を取得
dt = datetime.datetime.now()
date = dt.strftime("%Y-%m-%d")
# isdate = date.isnumeric
# # CSVファイルの存在確認　なければ作成
# logfile = Path('./log/' + date + '.csv')
# logfile.touch(exist_ok=True)
# print(logfile)

# wb = Workbook()
# ws1 = wb.create_sheet('NewSheet1', 0)
# ws2 = wb.create_sheet('addSheet2', 1)
# ws1.sheet_properties.tabColor = "1072BA"
# wb.save('Sample.xlsx')
path = 'Sample.xlsx'
wb = openpyxl.load_workbook(path)
ws = wb.active
ws['A1'] = 'this is test1'
ws.title = "test_sheet"
wb.save('Sample.xlsx')


# # ディレクトリ名を年月日にする
# DATE_DIR = './log/' + date
# # print(DATE_DIR)
# if os.path.isdir(DATE_DIR) == False :
#     # print('make new directory')
#     os.mkdir(DATE_DIR)
# # else:
# #     print('directory already exist')
# print(f'getpreferredencoding: {locale.getpreferredencoding()}')


# # logディレクトリ以下のファイルを読み込む
# # files = glob.glob(os.path.join(DATE_DIR,'*.db'))
# f = open('doorsingleKD.db')
# data = f.read()
# print(data)
# print(type(data))
# print(f'globed files: {f}')

# print(type(sys.argv))
# for x in sys.argv:
#     print(x)
#     print(type(x))
