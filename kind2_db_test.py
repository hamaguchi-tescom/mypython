# coding: utf-8
import sys
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


db_path = "./doorsingleKD.db"
time = sys.argv[1]

conn = sqlite3.connect(db_path)
sql = "SELECT add_time, equipment_name, work_number, recognition_name,temperature_val FROM recognition_record_view where add_time like ? order by add_time"

# 書き込み用ファイルを作成
FILL_TITLE = PatternFill(patternType='solid', fgColor='B0E0E6')
wb = Workbook()
ws = wb.active
y = 1

# 1行目にタイトルを設定
ws.cell(row=y, column=1).value = "add_time"
ws.cell(row=y, column=2).value = "equipment_name"
ws.cell(row=y, column=3).value = "work_number"
ws.cell(row=y, column=4).value = "recognition_name"
ws.cell(row=y, column=5).value = "temperature_val"

# 1行目だけ背景色を設定
for rows in ws['A1':'E1']:
    print(f'rows : {rows}')
    for cell in rows:
        print(f'cell : {cell}')
        cell.fill = FILL_TITLE

# ファイルから対象のデータを書き込む
for row in conn.execute(sql, (time+'%',)):
    y += 1
    ws.cell(row=y, column=1).value = row[0]
    ws.cell(row=y, column=2).value = row[1]
    ws.cell(row=y, column=3).value = row[2]
    ws.cell(row=y, column=4).value = row[3]
    ws.cell(row=y, column=5).value = row[4]

    for rows in ws.columns:
        max_length = 0
        for cell in rows:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[get_column_letter(
                cell.column)].width = adjusted_width

    wb.save(f'{time}.xlsx')
